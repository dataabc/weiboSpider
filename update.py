# excel_u.py
# coding=utf-8
# 导入相应模块
import datetime
import re

import easyquotation
import matplotlib
import numpy as np
import openpyxl
import requests
from bs4 import BeautifulSoup

# 处理乱码
matplotlib.rcParams['font.sans-serif'] = ['SimHei']
matplotlib.rcParams['font.family'] = 'sans-serif'
matplotlib.rcParams['axes.unicode_minus'] = False


# 查询股票信息
def get_stock(code):
    quotation = easyquotation.use('qq')  # 新浪 ['sina'] 腾讯 ['tencent', 'qq']
    return quotation.real(code)[code]  # 支持直接指定前缀，如 'sh000001'


# 每页最多50条数据
def get_html(code, start_date, end_date, page=1, per=10):
    url = 'http://fund.eastmoney.com/f10/F10DataApi.aspx?type=lsjz&code={0}&page={1}&sdate={2}&edate={3}&per={4}'.format(
        code, page, start_date, end_date, per)
    rsp = requests.get(url)
    html = rsp.text
    return html


def get_fund(code, start_date, end_date, page=1, per=20):
    # 获取html
    html = get_html(code, start_date, end_date, page, per)
    soup = BeautifulSoup(html, 'html.parser')
    # 获取总页数
    pattern = re.compile('pages:(.*),')
    result = re.search(pattern, html).group(1)
    total_page = int(result)
    # 获取表头信息
    heads = []
    for head in soup.findAll("th"):
        heads.append(head.contents[0])

    # 数据存取列表
    records = []
    # 获取每一页的数据
    current_page = 1
    while current_page <= total_page:
        html = get_html(code, start_date, end_date, current_page, per)
        soup = BeautifulSoup(html, 'html.parser')
        # 获取数据
        for row in soup.findAll("tbody")[0].findAll("tr"):
            row_records = []
            for record in row.findAll('td'):
                val = record.contents
                # 处理空值
                if not val:
                    row_records.append(np.nan)
                else:
                    row_records.append(val[0])
            # 记录数据
            # print (row_records[0] , row_records[1])
            records.append(row_records)
        # 下一页
        current_page = current_page + 1

    return records


# 获取前五天的时间
def getLastDay():
    today = datetime.date.today()
    oneday = datetime.timedelta(days=5)
    yesterday = today - oneday
    return yesterday


# 获取今天的时间
def getToday():
    return datetime.date.today()


# 场外基金
def run(code, row):
    start_date = getLastDay()
    end_date = getToday()

    print(code, 'begin==========>')

    records = get_fund(code, start_date, end_date)
    # 最新的净值记录
    record = records[0]
    # print(record)
    # 基金代码
    # table.cell(row,1,code)
    # 最新净值日期
    net_date = record[0]
    # 最新单位净值
    net_value = record[1]

    update_excel(net_date, net_value, row)
    print(code, 'end==========>')


# 场内基金或股票
def run_stock(code, row):
    print(code, 'begin==========>')
    data = get_stock(code)
    # print(data)
    # 最新净值日期
    # net_date = data['date']
    dt = data['datetime']
    # 格式化日期
    net_date = dt.strftime('%Y-%m-%d')
    # 最新单位净值
    net_value = data['now']
    update_excel(net_date, net_value, row)
    print(code, 'end==========>')


# 更新Excel
def update_excel(net_date, net_value, row):
    # 加载指定Excel
    data = openpyxl.load_workbook('AssetAllocation.xlsx')
    # 取第二张表
    table = data['明细']
    # 输出表名
    # print(table.title)
    table.cell(row, 11, net_date)
    table.cell(row, 12, net_value)
    data.save('AssetAllocation.xlsx')


def grid(code, row):
    # 加载指定Excel
    data = openpyxl.load_workbook('tt.xlsx')
    # oad_workbook(file, read_only=True, data_only=True)
    # 取第一张表
    table = data.worksheets[0]
    # 输出表名
    print(table.title)
    stock = get_stock(code)
    net_value = stock['now']
    print(net_value)
    table.cell(row, 2, net_value)
    data.save('tt.xlsx')


# 更新实验账户
def grid_1(code, cost, amount, dest):
    stock = get_stock(code)
    # print(stock)
    open = stock['open']
    now = stock['now']
    name = stock['name']
    profit = now - cost
    rate = profit / cost
    destRate = (dest - now) / now
    print(name,
          '成本价: {:.3f}'.format(cost),
          '现价: {:.3f}'.format(now),
          '涨跌幅: {:.2%}'.format(rate),
          '盈利: {:.2f}'.format(profit * amount),
          '目标价: {:.3f}'.format(dest),
          '目标涨跌幅: {:.2%}'.format(destRate))
    return profit * amount


def grid_notice(code, dest):
    stock = get_stock(code)
    # print(stock)
    now = stock['now']
    name = stock['name']
    if now >= dest:
        s = '叮咚！【' + name + '】达到目标价了，请及时操作哦~' + ' 现价: {:.3f}'.format(now)
        return s
    else:
        return '1'
